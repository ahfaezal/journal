import logging
import re
from collections import Counter
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[3]
UPLOAD_ROOT = PROJECT_ROOT / "storage" / "uploads"

logger = logging.getLogger(__name__)

YEAR_PATTERN = re.compile(r"\b(19|20)\d{2}[a-z]?\b")
NOISE_NAME_TOKENS = {
    "abdul",
    "abu",
    "al",
    "and",
    "b",
    "binti",
    "bin",
    "bt",
    "dan",
    "de",
    "dr",
    "et",
    "ibn",
    "mohamad",
    "mohd",
    "muhammad",
    "nur",
    "van",
    "wan",
}


def build_citation_map(
    project_id: str,
    parsed_thesis: dict[str, Any],
    upload_metadata: dict[str, dict[str, str]],
) -> dict[str, Any]:
    parsed_citations = expand_parsed_citations(parsed_thesis.get("citations", []))
    parsed_headings = parsed_thesis.get("headings", [])
    citation_texts = [
        str(citation.get("citation", ""))
        for citation in parsed_citations
        if isinstance(citation, dict) and citation.get("citation")
    ]
    citation_counts = Counter(citation_texts)

    mfl_files = find_mfl_files(project_id, upload_metadata)
    logger.info("MFL file detected: project_id=%s count=%s files=%s", project_id, len(mfl_files), [path.name for path in mfl_files])
    mfl_entries = parse_mfl_files(mfl_files)
    logger.info("MFL entries parsed: project_id=%s count=%s", project_id, len(mfl_entries))
    logger.info("citations detected: project_id=%s count=%s", project_id, len(parsed_citations))

    citation_rows = []
    matched_citations = 0
    unmatched_citations = 0
    issue_counts: Counter[str] = Counter()

    for citation in parsed_citations:
        if not isinstance(citation, dict):
            continue

        citation_text = str(citation.get("citation", ""))
        detected_author = str(citation.get("author", ""))
        detected_year = str(citation.get("year", ""))
        source_file = str(citation.get("source_file", "Unknown file"))
        source_position = int(citation.get("position", 0) or 0)
        source_section = detect_source_section(parsed_headings, source_file, source_position)
        duplicate_count = citation_counts.get(citation_text, 0)
        match = find_mfl_match(detected_author, detected_year, mfl_entries)
        is_matched = match is not None

        if is_matched:
            matched_citations += 1
        else:
            unmatched_citations += 1

        issue = build_issue(
            mfl_available=bool(mfl_files),
            references_count=len(mfl_entries),
            matched=is_matched,
            detected_author=detected_author,
            detected_year=detected_year,
            duplicate_count=duplicate_count,
        )
        if issue != "No issue detected":
            for part in issue.split("; "):
                issue_counts[part] += 1

        citation_rows.append(
            {
                "citation_text": citation_text,
                "detected_author": detected_author,
                "detected_year": detected_year,
                "source_file": source_file,
                "source_section": source_section,
                "mfl_status": "matched" if is_matched else "unmatched",
                "matched_reference": match.get("raw_reference", "") if match else "",
                "matched_title": match.get("title", "") if match else "",
                "issue": issue,
            }
        )

    match_rate = round((matched_citations / len(citation_rows)) * 100, 1) if citation_rows else 0
    logger.info("citation matches found: project_id=%s matched=%s total=%s rate=%s", project_id, matched_citations, len(citation_rows), match_rate)

    return {
        "project_id": project_id,
        "status": "mapped",
        "mfl_available": bool(mfl_files),
        "mfl_files": [path.name for path in mfl_files],
        "references_count": len(mfl_entries),
        "total_citations": len(citation_rows),
        "unique_citations": len(citation_counts),
        "matched_citations": matched_citations,
        "unmatched_citations": unmatched_citations,
        "duplicate_citations": sum(count - 1 for count in citation_counts.values() if count > 1),
        "match_rate": match_rate,
        "mfl_match_status": f"{match_rate}% matched" if mfl_entries else "MFL not parsed",
        "issues": dict(issue_counts),
        "mfl_references": mfl_entries,
        "citations": citation_rows,
    }


def find_mfl_files(project_id: str, upload_metadata: dict[str, dict[str, str]]) -> list[Path]:
    upload_dir = UPLOAD_ROOT / project_id
    candidates: list[Path] = []

    for filename, metadata in upload_metadata.items():
        file_type = str(metadata.get("file_type", "")).lower()
        chapter_label = str(metadata.get("chapter_label", "")).lower()
        stored_filename = str(metadata.get("stored_filename") or filename)
        metadata_path = metadata.get("file_path")
        possible_path = Path(metadata_path) if metadata_path else upload_dir / stored_filename
        if ("mfl" in file_type or "reference" in file_type or "mfl" in chapter_label) and possible_path.exists():
            candidates.append(possible_path)

    if upload_dir.exists():
        for file_path in upload_dir.iterdir():
            if not file_path.is_file() or file_path.name.startswith("."):
                continue
            lowered = file_path.name.lower()
            if file_path.suffix.lower() in {".xlsx", ".xlsm", ".csv"} and (
                "mfl" in lowered or "master" in lowered or "reference" in lowered
            ):
                candidates.append(file_path)

    return deduplicate_paths(candidates)


def parse_mfl_files(file_paths: list[Path]) -> list[dict[str, str]]:
    entries: list[dict[str, str]] = []
    for file_path in file_paths:
        if file_path.suffix.lower() in {".xlsx", ".xlsm"}:
            entries.extend(parse_mfl_xlsx(file_path))
        elif file_path.suffix.lower() == ".csv":
            entries.extend(parse_mfl_csv(file_path))

    return deduplicate_entries(entries)


def parse_mfl_xlsx(file_path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl is not installed; cannot parse MFL XLSX: %s", file_path)
        return []

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as error:
        logger.warning("Unable to read MFL XLSX %s: %s", file_path, error)
        return []

    entries: list[dict[str, str]] = []
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_index, column_map = detect_header_and_columns(rows)
        data_rows = rows[header_index + 1 :] if column_map else rows
        for row in data_rows:
            entry = build_mfl_entry(row, column_map, file_path.name)
            if entry:
                entries.append(entry)

    workbook.close()
    return entries


def parse_mfl_csv(file_path: Path) -> list[dict[str, str]]:
    import csv

    try:
        with file_path.open("r", encoding="utf-8-sig", newline="") as csv_file:
            rows = list(csv.reader(csv_file))
    except UnicodeDecodeError:
        with file_path.open("r", encoding="latin-1", newline="") as csv_file:
            rows = list(csv.reader(csv_file))
    except Exception as error:
        logger.warning("Unable to read MFL CSV %s: %s", file_path, error)
        return []

    header_index, column_map = detect_header_and_columns(rows)
    data_rows = rows[header_index + 1 :] if column_map else rows
    return [entry for row in data_rows if (entry := build_mfl_entry(row, column_map, file_path.name))]


def detect_header_and_columns(rows: list[Any]) -> tuple[int, dict[str, int]]:
    best_index = 0
    best_score = 0
    best_map: dict[str, int] = {}
    for index, row in enumerate(rows[:10]):
        values = [normalize_cell(value) for value in row]
        column_map = map_columns(values)
        score = len(column_map)
        if score > best_score:
            best_index = index
            best_score = score
            best_map = column_map

    return best_index, best_map


def map_columns(headers: list[str]) -> dict[str, int]:
    column_map: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = normalize_key(header)
        if not normalized:
            continue
        if "author" in normalized or "pengarang" in normalized or "nama" == normalized:
            column_map.setdefault("author", index)
        elif "year" in normalized or "tahun" in normalized:
            column_map.setdefault("year", index)
        elif "title" in normalized or "tajuk" in normalized:
            column_map.setdefault("title", index)
        elif "reference" in normalized or "rujukan" in normalized or "apa" in normalized:
            column_map.setdefault("reference", index)
        elif "source" in normalized or "journal" in normalized or "publisher" in normalized:
            column_map.setdefault("source", index)

    return column_map


def build_mfl_entry(row: Any, column_map: dict[str, int], source_file: str) -> dict[str, str] | None:
    values = [normalize_cell(value) for value in row]
    if not any(values):
        return None

    raw_reference = pick_column(values, column_map, "reference")
    author_text = pick_column(values, column_map, "author")
    year = pick_column(values, column_map, "year")
    title = pick_column(values, column_map, "title")
    source = pick_column(values, column_map, "source")

    if not raw_reference:
        raw_reference = infer_raw_reference(values)
    if not year:
        year = infer_year(raw_reference or " ".join(values))
    if not author_text:
        author_text = infer_author(raw_reference, values)
    if not title:
        title = infer_title(raw_reference, values)

    if not raw_reference and not (author_text and year):
        return None

    return {
        "source_file": source_file,
        "author_text": author_text,
        "normalized_author": normalize_author(author_text),
        "year": year[:4] if year else "",
        "title": title,
        "source": source,
        "raw_reference": raw_reference or " ".join(value for value in values if value),
    }


def expand_parsed_citations(parsed_citations: Any) -> list[dict[str, Any]]:
    expanded: list[dict[str, Any]] = []
    if not isinstance(parsed_citations, list):
        return expanded

    for citation in parsed_citations:
        if not isinstance(citation, dict):
            continue
        citation_text = str(citation.get("citation", ""))
        parts = split_citation_text(citation_text)
        if not parts:
            expanded.append(citation)
            continue
        for part in parts:
            author, year = parse_citation_part(part)
            cloned = dict(citation)
            cloned["citation"] = part
            cloned["author"] = author or citation.get("author", "")
            cloned["year"] = year or citation.get("year", "")
            expanded.append(cloned)

    return expanded


def split_citation_text(citation_text: str) -> list[str]:
    cleaned = citation_text.strip()
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = cleaned[1:-1]
    if ";" not in cleaned:
        return [citation_text] if YEAR_PATTERN.search(citation_text) else []

    parts = []
    for part in cleaned.split(";"):
        normalized = part.strip()
        if normalized and YEAR_PATTERN.search(normalized):
            parts.append(f"({normalized})")
    return parts


def parse_citation_part(citation_text: str) -> tuple[str, str]:
    text = citation_text.strip("() ")
    year_match = YEAR_PATTERN.search(text)
    year = year_match.group(0)[:4] if year_match else ""
    author = text[: year_match.start()].strip(" ,") if year_match else text
    return author, year


def find_mfl_match(author: str, year: str, entries: list[dict[str, str]]) -> dict[str, str] | None:
    if not year:
        return None

    citation_tokens = author_tokens(author)
    if not citation_tokens:
        return None

    for entry in entries:
        if str(entry.get("year", ""))[:4] != year[:4]:
            continue
        entry_tokens = author_tokens(entry.get("author_text", "") or entry.get("raw_reference", ""))
        if citation_tokens & entry_tokens:
            return entry

    return None


def author_tokens(author_text: str) -> set[str]:
    normalized = normalize_author(author_text.replace("&", " ").replace("et al.", " "))
    tokens = {
        token
        for token in normalized.split()
        if len(token) > 2 and token not in NOISE_NAME_TOKENS
    }
    if tokens:
        return tokens

    fallback = [token for token in normalized.split() if len(token) > 1]
    return set(fallback[:1])


def normalize_author(author_text: str) -> str:
    text = author_text.lower()
    text = re.sub(r"\bet\s+al\.?", " ", text)
    text = re.sub(r"[^a-zÀ-ÿ0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def detect_source_section(
    headings: object,
    source_file: str,
    source_position: int,
) -> str:
    if not isinstance(headings, list):
        return "Unknown section"

    matched_heading = None
    for heading in headings:
        if not isinstance(heading, dict):
            continue

        heading_file = str(heading.get("source_file", ""))
        heading_position = int(heading.get("position", 0) or 0)
        if heading_file == source_file and heading_position <= source_position:
            matched_heading = str(heading.get("text", "Unknown section"))

    return matched_heading or "Unknown section"


def build_issue(
    mfl_available: bool,
    references_count: int,
    matched: bool,
    detected_author: str,
    detected_year: str,
    duplicate_count: int,
) -> str:
    issues = []
    if not mfl_available:
        issues.append("MFL not uploaded")
    elif references_count == 0:
        issues.append("MFL could not be parsed")
    if not detected_year:
        issues.append("citation year missing")
    if not detected_author:
        issues.append("citation author missing")
    if mfl_available and references_count and not matched:
        issues.append("no matching MFL reference")
    if duplicate_count > 1:
        issues.append("duplicate citation")

    return "; ".join(issues) if issues else "No issue detected"


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def pick_column(values: list[str], column_map: dict[str, int], key: str) -> str:
    index = column_map.get(key)
    if index is None or index >= len(values):
        return ""
    return values[index]


def infer_raw_reference(values: list[str]) -> str:
    candidates = [value for value in values if YEAR_PATTERN.search(value) and len(value) > 30]
    if candidates:
        return max(candidates, key=len)
    return " ".join(value for value in values if value)


def infer_year(text: str) -> str:
    match = YEAR_PATTERN.search(text)
    return match.group(0)[:4] if match else ""


def infer_author(raw_reference: str, values: list[str]) -> str:
    text = raw_reference or " ".join(values)
    if not text:
        return ""
    year_match = YEAR_PATTERN.search(text)
    before_year = text[: year_match.start()] if year_match else text
    return before_year.split(".")[0].strip(" ,")


def infer_title(raw_reference: str, values: list[str]) -> str:
    if raw_reference:
        parts = [part.strip() for part in raw_reference.split(".") if part.strip()]
        if len(parts) >= 2:
            return parts[1]
    long_values = [value for value in values if len(value) > 15 and not YEAR_PATTERN.fullmatch(value)]
    return long_values[0] if long_values else ""


def deduplicate_entries(entries: list[dict[str, str]]) -> list[dict[str, str]]:
    seen = set()
    deduplicated = []
    for entry in entries:
        marker = (
            entry.get("normalized_author", ""),
            entry.get("year", ""),
            entry.get("title", ""),
            entry.get("raw_reference", ""),
        )
        if marker in seen:
            continue
        seen.add(marker)
        deduplicated.append(entry)
    return deduplicated


def deduplicate_paths(paths: list[Path]) -> list[Path]:
    seen = set()
    deduplicated = []
    for path in paths:
        marker = str(path.resolve())
        if marker in seen:
            continue
        seen.add(marker)
        deduplicated.append(path)
    return deduplicated
